#!/usr/bin/env python3
"""
OrbitSight Challenge — Evaluation Script
=========================================
Computes detection metrics by comparing predicted bounding boxes against
ground truth for the OrbitSight neuromorphic dataset.

Ground truth format  (Dataset_GT/*_bb_windows_40ms.txt):
  window_start_timestamp_us  window_end_timestamp_us  center_x  center_y  width  height

Prediction format (pred_dir/*_bb_windows_40ms.txt):
  window_start_timestamp_us  window_end_timestamp_us  center_x  center_y  width  height  confidence
  (confidence is optional — defaults to 1.0 if omitted)

Metrics reported:
  Precision, Recall, F1  @ IoU 0.5   (per sequence + overall)
  AP @ IoU 0.5                        (per sequence)
  mAP @ IoU 0.5                       (overall)

Usage (training + testing split):
  python3 evaluate.py \
    --train-gt-dir ../OrbitSight_Dataset/Training_sets \
    --train-pred-dir ../predictions/training \
    --test-gt-dir  ../OrbitSight_Dataset/Testing_sets \
    --test-pred-dir ../predictions/testing

Usage (single directory):
  python3 evaluate.py --gt-dir ../Dataset_GT --pred-dir ../predictions

Outputs Evaluation_Metric.xlsx in the current directory.
Expected file naming: prediction files must match GT filenames exactly.
"""

import argparse
import csv
import os

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from tabulate import tabulate


IOU_THRESHOLD = 0.5

# ── Colour palette ────────────────────────────────────────────────────────────
_BLUE_HEADER  = "FF1F3864"   # dark navy  — header row
_BLUE_TRAIN   = "FFD6E4F0"   # light blue — training rows
_GREEN_TEST   = "FFE2EFDA"   # light green — testing rows
_YELLOW_TOTAL = "FFFFF2CC"   # light yellow — total row
_WHITE        = "FFFFFFFF"


# ─── Box helpers ─────────────────────────────────────────────────────────────

def cx_cy_wh_to_xyxy(cx, cy, w, h):
    x1 = cx - (w - 1) / 2
    y1 = cy - (h - 1) / 2
    x2 = x1 + w - 1
    y2 = y1 + h - 1
    return x1, y1, x2, y2


def iou(box_a, box_b):
    """IoU between two boxes given as (cx, cy, w, h)."""
    ax1, ay1, ax2, ay2 = cx_cy_wh_to_xyxy(*box_a)
    bx1, by1, bx2, by2 = cx_cy_wh_to_xyxy(*box_b)

    ix1, iy1 = max(ax1, bx1), max(ay1, by1)
    ix2, iy2 = min(ax2, bx2), min(ay2, by2)
    inter = max(0, ix2 - ix1 + 1) * max(0, iy2 - iy1 + 1)

    area_a = (ax2 - ax1 + 1) * (ay2 - ay1 + 1)
    area_b = (bx2 - bx1 + 1) * (by2 - by1 + 1)
    union  = area_a + area_b - inter
    return inter / union if union > 0 else 0.0


# ─── File loading ─────────────────────────────────────────────────────────────

def load_gt(path):
    """Returns list of (window_start_us, window_end_us, cx, cy, w, h)."""
    rows = []
    with open(path, newline="") as f:
        reader = csv.DictReader(f, delimiter="\t")
        for row in reader:
            rows.append((
                int(row["window_start_timestamp_us"]),
                int(row["window_end_timestamp_us"]),
                int(row["center_x"]),
                int(row["center_y"]),
                int(row["width"]),
                int(row["height"]),
            ))
    return rows


def load_pred(path):
    """Returns list of (window_start_us, window_end_us, cx, cy, w, h, confidence)."""
    rows = []
    with open(path, newline="") as f:
        reader = csv.DictReader(f, delimiter="\t")
        for row in reader:
            conf = float(row["confidence"]) if "confidence" in row else 1.0
            rows.append((
                int(row["window_start_timestamp_us"]),
                int(row["window_end_timestamp_us"]),
                int(row["center_x"]),
                int(row["center_y"]),
                int(row["width"]),
                int(row["height"]),
                conf,
            ))
    rows.sort(key=lambda r: r[6], reverse=True)
    return rows


# ─── Matching ─────────────────────────────────────────────────────────────────

def windows_overlap(ws_a, we_a, ws_b, we_b):
    return ws_a < we_b and we_a > ws_b


def match_predictions(gt_list, pred_list, iou_thresh=IOU_THRESHOLD):
    gt_matched = [False] * len(gt_list)
    tp = []
    fp = []

    for pred in pred_list:
        ws_p, we_p, cx_p, cy_p, w_p, h_p, _ = pred
        best_iou  = 0.0
        best_idx  = -1

        for j, gt in enumerate(gt_list):
            if gt_matched[j]:
                continue
            ws_g, we_g, cx_g, cy_g, w_g, h_g = gt
            if not windows_overlap(ws_p, we_p, ws_g, we_g):
                continue
            score = iou((cx_p, cy_p, w_p, h_p), (cx_g, cy_g, w_g, h_g))
            if score > best_iou:
                best_iou = score
                best_idx = j

        if best_iou >= iou_thresh and best_idx >= 0:
            tp.append(1)
            fp.append(0)
            gt_matched[best_idx] = True
        else:
            tp.append(0)
            fp.append(1)

    return np.array(tp), np.array(fp)


# ─── Metrics ─────────────────────────────────────────────────────────────────

def compute_ap(tp, fp, n_gt):
    if n_gt == 0:
        return float("nan")

    cum_tp = np.cumsum(tp)
    cum_fp = np.cumsum(fp)

    recalls    = cum_tp / n_gt
    precisions = cum_tp / (cum_tp + cum_fp + 1e-9)

    recalls    = np.concatenate([[0.0], recalls,    [recalls[-1]  if len(recalls)  else 0.0]])
    precisions = np.concatenate([[1.0], precisions, [0.0]])

    for i in range(len(precisions) - 2, -1, -1):
        precisions[i] = max(precisions[i], precisions[i + 1])

    idx = np.where(recalls[1:] != recalls[:-1])[0]
    ap  = np.sum((recalls[idx + 1] - recalls[idx]) * precisions[idx + 1])
    return float(ap)


def compute_prf1(tp_total, fp_total, fn_total):
    precision = tp_total / (tp_total + fp_total) if (tp_total + fp_total) > 0 else 0.0
    recall    = tp_total / (tp_total + fn_total) if (tp_total + fn_total) > 0 else 0.0
    f1        = (2 * precision * recall / (precision + recall)
                 if (precision + recall) > 0 else 0.0)
    return precision, recall, f1


# ─── Evaluate one directory ───────────────────────────────────────────────────

def evaluate_dir(gt_dir, pred_dir, iou_thresh, label):
    """
    Evaluate all sequences in gt_dir against pred_dir.
    Returns list of dicts, one per sequence.
    """
    gt_files = sorted(f for f in os.listdir(gt_dir) if f.endswith("_bb_windows_40ms.txt"))
    results = []

    for fname in gt_files:
        seq_name  = fname.replace("_bb_windows_40ms.txt", "")
        gt_path   = os.path.join(gt_dir,   fname)
        pred_path = os.path.join(pred_dir, fname)

        if not os.path.exists(pred_path):
            print(f"[WARN] Missing prediction for: {fname} — skipping")
            continue

        gt_list   = load_gt(gt_path)
        pred_list = load_pred(pred_path)
        n_gt   = len(gt_list)
        n_pred = len(pred_list)

        if n_gt == 0 and n_pred == 0:
            results.append(dict(
                label=label, seq=seq_name, n_gt=0, n_pred=0,
                prec=None, rec=None, f1=None, ap=None,
                tp=0, fp=0, fn=0,
            ))
            continue

        tp_arr, fp_arr = match_predictions(gt_list, pred_list, iou_thresh)
        tp = int(tp_arr.sum())
        fp = int(fp_arr.sum())
        fn = n_gt - tp

        prec, rec, f1 = compute_prf1(tp, fp, fn)
        ap = compute_ap(tp_arr, fp_arr, n_gt)

        results.append(dict(
            label=label, seq=seq_name, n_gt=n_gt, n_pred=n_pred,
            prec=prec, rec=rec, f1=f1, ap=ap,
            tp=tp, fp=fp, fn=fn,
        ))

    return results


# ─── Excel output ─────────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    s = Side(style="thin", color="FF999999")
    return Border(left=s, right=s, top=s, bottom=s)

def _fmt(v, decimals=4):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "—"
    return round(float(v), decimals)


def write_excel(all_results, excel_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Evaluation"

    headers = ["Type", "Sequence", "GT Boxes", "Pred Boxes",
               "Precision", "Recall", "F1 Score", "AP @ IoU 0.5",
               "TP", "FP", "FN"]

    # ── header row ──
    header_font = Font(name="Calibri", bold=True, color="FFFFFFFF", size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font    = header_font
        cell.fill    = _fill(_BLUE_HEADER)
        cell.border  = _border()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── data rows ──
    all_tp = all_fp = all_fn = 0
    all_ap = []
    row_idx = 2

    for r in all_results:
        fill_color = _BLUE_TRAIN if r["label"] == "Training" else _GREEN_TEST
        row_font   = Font(name="Calibri", size=11)

        values = [
            r["label"],
            r["seq"],
            r["n_gt"],
            r["n_pred"],
            _fmt(r["prec"]),
            _fmt(r["rec"]),
            _fmt(r["f1"]),
            _fmt(r["ap"]),
            r["tp"],
            r["fp"],
            r["fn"],
        ]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.font      = row_font
            cell.fill      = _fill(fill_color)
            cell.border    = _border()
            cell.alignment = Alignment(horizontal="center" if col != 2 else "left",
                                       vertical="center")

        all_tp += r["tp"]
        all_fp += r["fp"]
        all_fn += r["fn"]
        if r["ap"] is not None and not np.isnan(r["ap"]):
            all_ap.append(r["ap"])

        row_idx += 1

    # ── total / average row ──
    overall_prec, overall_rec, overall_f1 = compute_prf1(all_tp, all_fp, all_fn)
    map50 = float(np.mean(all_ap)) if all_ap else float("nan")

    total_values = [
        "Overall",
        f"Average ({len(all_results)} sequences)",
        "",
        "",
        _fmt(overall_prec),
        _fmt(overall_rec),
        _fmt(overall_f1),
        _fmt(map50),
        all_tp,
        all_fp,
        all_fn,
    ]
    total_font = Font(name="Calibri", bold=True, size=11)
    for col, v in enumerate(total_values, 1):
        cell = ws.cell(row=row_idx, column=col, value=v)
        cell.font      = total_font
        cell.fill      = _fill(_YELLOW_TOTAL)
        cell.border    = _border()
        cell.alignment = Alignment(horizontal="center" if col != 2 else "left",
                                   vertical="center")

    # ── column widths ──
    col_widths = [12, 52, 10, 12, 11, 9, 11, 14, 7, 7, 7]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    wb.save(excel_path)
    print(f"[INFO] Excel saved → {excel_path}")


# ─── Console print ────────────────────────────────────────────────────────────

def print_results(all_results):
    table_rows = []
    all_tp = all_fp = all_fn = 0
    all_ap = []

    for r in all_results:
        prec = f"{r['prec']:.4f}" if r["prec"] is not None else "—"
        rec  = f"{r['rec']:.4f}"  if r["rec"]  is not None else "—"
        f1   = f"{r['f1']:.4f}"   if r["f1"]   is not None else "—"
        ap   = f"{r['ap']:.4f}"   if (r["ap"] is not None and not np.isnan(r["ap"])) else "—"
        table_rows.append([r["label"], r["seq"], r["n_gt"], r["n_pred"], prec, rec, f1, ap])
        all_tp += r["tp"]; all_fp += r["fp"]; all_fn += r["fn"]
        if r["ap"] is not None and not np.isnan(r["ap"]):
            all_ap.append(r["ap"])

    headers = ["Type", "Sequence", "GT Boxes", "Pred Boxes", "Precision", "Recall", "F1", "AP@0.5"]
    print("\n" + "=" * 120)
    print("  OrbitSight Detection Evaluation — Per Sequence")
    print("=" * 120)
    print(tabulate(table_rows, headers=headers, tablefmt="github"))

    overall_prec, overall_rec, overall_f1 = compute_prf1(all_tp, all_fp, all_fn)
    map50 = float(np.mean(all_ap)) if all_ap else float("nan")

    summary = [
        ["mAP @ IoU 0.5",      f"{map50:.4f}"],
        ["Precision",           f"{overall_prec:.4f}"],
        ["Recall",              f"{overall_rec:.4f}"],
        ["F1 Score",            f"{overall_f1:.4f}"],
        ["Total TP",            all_tp],
        ["Total FP",            all_fp],
        ["Total FN (missed)",   all_fn],
    ]
    print("\n" + "=" * 50)
    print("  Overall Results")
    print("=" * 50)
    print(tabulate(summary, tablefmt="github"))
    print()


# ─── Main ─────────────────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(description="OrbitSight evaluation script")

    # Training set
    p.add_argument("--train-gt-dir",   default=None,
                   help="GT directory for training sequences")
    p.add_argument("--train-pred-dir", default=None,
                   help="Prediction directory for training sequences")

    # Testing set
    p.add_argument("--test-gt-dir",    default=None,
                   help="GT directory for testing sequences")
    p.add_argument("--test-pred-dir",  default=None,
                   help="Prediction directory for testing sequences")

    # Single-directory fallback
    p.add_argument("--gt-dir",   default=None,
                   help="GT directory (single-split mode)")
    p.add_argument("--pred-dir", default=None,
                   help="Prediction directory (single-split mode)")

    p.add_argument("--iou", type=float, default=IOU_THRESHOLD,
                   help=f"IoU threshold (default: {IOU_THRESHOLD})")
    p.add_argument("--excel-out", default="Evaluation_Metric.xlsx",
                   help="Output Excel filename (default: Evaluation_Metric.xlsx)")
    return p.parse_args()


def main():
    args = parse_args()

    all_results = []

    if args.train_gt_dir and args.train_pred_dir:
        print(f"[INFO] Evaluating training sequences in: {args.train_gt_dir}")
        all_results += evaluate_dir(args.train_gt_dir, args.train_pred_dir, args.iou, "Training")

    if args.test_gt_dir and args.test_pred_dir:
        print(f"[INFO] Evaluating testing sequences in:  {args.test_gt_dir}")
        all_results += evaluate_dir(args.test_gt_dir, args.test_pred_dir, args.iou, "Testing")

    if not all_results:
        if args.gt_dir and args.pred_dir:
            print(f"[INFO] Evaluating sequences in: {args.gt_dir}")
            all_results += evaluate_dir(args.gt_dir, args.pred_dir, args.iou, "")
        else:
            print("[ERROR] Provide --train-gt-dir/--train-pred-dir and/or "
                  "--test-gt-dir/--test-pred-dir, or --gt-dir/--pred-dir.")
            return

    if not all_results:
        print("[ERROR] No sequences evaluated.")
        return

    print_results(all_results)
    write_excel(all_results, args.excel_out)


if __name__ == "__main__":
    main()
